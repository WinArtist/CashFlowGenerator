"""季度合并服务 - 将三个月的Sheet合并成季度Sheet"""

import openpyxl
from openpyxl.utils.cell import get_column_letter
from typing import List


def merge_quarter_sheets(wb, source_sheets: List[str], target_sheet_name: str, bank_name: str, quarter: str) -> bool:
    """
    将三个月的Sheet合并成季度Sheet
    
    Args:
        wb: 工作簿对象
        source_sheets: 源Sheet名称列表（三个月的Sheet）
        target_sheet_name: 目标Sheet名称（季度）
        bank_name: 银行名称
        quarter: 季度（Q1/Q2/Q3/Q4）
    
    Returns:
        是否成功
    """
    try:
        # 创建一个新的Sheet作为季度合并结果
        # 先获取第一个Sheet作为模板
        template_ws = wb[source_sheets[0]]
        
        # 如果目标Sheet已存在，删除
        if target_sheet_name in wb.sheetnames:
            wb.remove(wb[target_sheet_name])
        
        # 复制第一个Sheet作为模板
        ws = wb.copy_worksheet(template_ws)
        ws.title = target_sheet_name
        
        # 清空数据行（保留结构和公式）
        # 获取数据行范围（从第5行开始）
        max_row = ws.max_row
        for row in range(5, max_row + 1):
            # 清空数据列（J到Q列）
            for col in range(10, 18):  # J到Q列
                ws.cell(row=row, column=col).value = None
            
            # 清空分类列（V到BF列）
            for col in range(22, 58):  # V到BF列
                ws.cell(row=row, column=col).value = None
        
        # 遍历三个源Sheet，累加数据到目标Sheet
        row_count = 0
        current_row = 5
        
        for sheet_name in source_sheets:
            src_ws = wb[sheet_name]
            src_max_row = src_ws.max_row
            
            # 遍历数据行（从第5行开始）
            for src_row in range(5, src_max_row + 1):
                # 检查是否有数据
                has_data = False
                for col in range(10, 18):  # J到Q列
                    val = src_ws.cell(row=src_row, column=col).value
                    if val and val != 0:
                        has_data = True
                        break
                if not has_data:
                    for col in range(22, 58):  # V到BF列
                        val = src_ws.cell(row=src_row, column=col).value
                        if val and val != 0:
                            has_data = True
                            break
                
                if not has_data:
                    continue
                
                # 复制数据行
                for col in range(1, ws.max_column + 1):
                    val = src_ws.cell(row=src_row, column=col).value
                    if val:
                        ws.cell(row=current_row, column=col, value=val)
                
                current_row += 1
                row_count += 1
        
        self.log_text.append(f'  合并了 {row_count} 行数据')
        
        # 更新期初余额（第4行）
        # 从第一个Sheet取期初余额
        first_sheet = wb[source_sheets[0]]
        ws.cell(row=4, column=18, value=first_sheet.cell(row=4, column=18).value)  # R4
        
        return True
        
    except Exception as e:
        print(f"❌ 季度合并失败: {e}")
        import traceback
        traceback.print_exc()
        return False