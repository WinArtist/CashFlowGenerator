# src/services/summary_service.py
"""汇总表服务 - 用公式引用已有Sheet的合计行数据"""

from decimal import Decimal
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import Config


class SummaryService:
    """汇总表服务 - 用公式引用已有Sheet的合计数据"""
    
    def __init__(self, config: Config):
        self.config = config
        
        # 汇总表需要引用公式的列（C列到BF列，即 3-58）
        self.FORMULA_COLUMNS = list(range(3, 59))
    
    def parse_sheet_name(self, sheet_name: str) -> Tuple[Optional[str], Optional[str]]:
        """
        解析Sheet名称，提取银行名称和月份
        
        示例:
            "上海银行_1月" -> ("上海银行", "1月")
            "招商银行_12月" -> ("招商银行", "12月")
        """
        pattern = r'^(.+)_(\d+月)$'
        match = re.match(pattern, sheet_name)
        if match:
            return match.group(1), match.group(2)
        return None, None
    
    def _month_to_number(self, month_str: str) -> Optional[int]:
        """将月份字符串转换为数字"""
        match = re.search(r'(\d+)', month_str)
        if match:
            num = int(match.group(1))
            if 1 <= num <= 12:
                return num
        return None
    
    def _find_summary_row(self, ws) -> Optional[int]:
        """
        查找Sheet的合计行（数据行之后的合计行）
        从最后一行往上找，找到有合计数据的行
        如果Sheet没有有效数据，返回 None
        """
        max_row = ws.max_row
        
        # ===== Sheet至少要有10行以上才可能有数据 =====
        if max_row < 10:
            return None
        
        # 从最后一行往上找
        for row in range(max_row, 4, -1):
            cell_c = ws.cell(row=row, column=3)
            cell_r = ws.cell(row=row, column=18)
            if cell_c.value is not None or cell_r.value is not None:
                # 进一步确认：检查该行是否有数据
                has_data = False
                for col in [3, 4, 5, 6, 7, 8, 9, 15, 16, 17, 22, 23]:
                    val = ws.cell(row=row, column=col).value
                    if val is not None and str(val).strip() != '':
                        has_data = True
                        break
                if has_data:
                    return row
        return None
    
    def generate_formula(self, sheet_name: str, summary_row: int, col: int) -> str:
        """生成引用公式"""
        col_letter = get_column_letter(col)
        if "'" in sheet_name:
            sheet_name = sheet_name.replace("'", "''")
        return f"='{sheet_name}'!{col_letter}{summary_row}"
    
    def get_sheets_from_workbook(self, wb) -> List[Dict]:
        """从工作簿中获取所有符合格式的Sheet"""
        sheets = []
        for sheet_name in wb.sheetnames:
            bank, month = self.parse_sheet_name(sheet_name)
            if bank and month:
                ws = wb[sheet_name]
                summary_row = self._find_summary_row(ws)
                if summary_row is not None:
                    sheets.append({
                        "bank": bank,
                        "month": month,
                        "sheet": sheet_name,
                        "summary_row": summary_row
                    })
                else:
                    print(f"  ⚠️ 跳过 {sheet_name} (未找到合计行)")
        
        sheets.sort(key=lambda x: (x["bank"], self._month_to_number(x["month"]) or 0))
        return sheets
    
    def _is_merged_cell(self, ws, row: int, col: int) -> bool:
        """检查单元格是否为合并单元格的一部分"""
        for merged_range in ws.merged_cells.ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                return True
        return False
    
    def fill_summary_sheet(self, template_path: str, file_paths: List[Path], classification_config) -> Optional[str]:
        """填充汇总表 - 用公式引用已有Sheet的合计行"""
        from openpyxl import load_workbook
        
        if not template_path:
            raise ValueError("模板文件路径未设置")
        
        template_path = Path(template_path)
        if not template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_path}")
        
        wb = load_workbook(template_path)
        
        # 获取所有符合格式的Sheet
        sheets = self.get_sheets_from_workbook(wb)
        
        if not sheets:
            print("❌ 没有找到符合条件的Sheet（格式: 银行_月份）")
            print("💡 请先运行'单文件生成'模式创建Sheet")
            return None
        
        print(f"\n📊 找到 {len(sheets)} 个有效Sheet:")
        for s in sheets:
            print(f"  {s['sheet']} (合计行: {s['summary_row']})")
        
        # 找到汇总表
        summary_sheet_name = None
        for sheet_name in wb.sheetnames:
            if '汇总表' in sheet_name:
                summary_sheet_name = sheet_name
                break
        
        if summary_sheet_name is None:
            print("⚠️ 未找到'汇总表'，将创建新Sheet")
            ws = wb.create_sheet("汇总表")
            ws.cell(row=1, column=1, value="行号")
            ws.cell(row=1, column=2, value="交易时间")
            ws.cell(row=2, column=1, value="银行")
            ws.cell(row=2, column=2, value="月份")
            summary_sheet_name = "汇总表"
        else:
            ws = wb[summary_sheet_name]
        
        # ===== 字体：微软雅黑 8号 =====
        font = Font(name='微软雅黑', size=8)
        number_format = '#,##0.00;-#,##0.00;-'
        
        # ===== 清空汇总表数据行（从第4行开始），跳过合并单元格 =====
        start_row = 4
        max_row = ws.max_row
        if max_row >= start_row:
            for row in range(start_row, max_row + 1):
                for col in range(1, ws.max_column + 1):
                    if not self._is_merged_cell(ws, row, col):
                        cell = ws.cell(row=row, column=col)
                        cell.value = None
                        cell.font = font
        
        # 为每个Sheet生成一行引用公式
        for idx, sheet_info in enumerate(sheets):
            current_row = start_row + idx
            sheet_name = sheet_info["sheet"]
            summary_row = sheet_info["summary_row"]
            
            # A列：银行名称
            if not self._is_merged_cell(ws, current_row, 1):
                cell = ws.cell(row=current_row, column=1, value=sheet_info["bank"])
                cell.font = font
            
            # B列：月份
            if not self._is_merged_cell(ws, current_row, 2):
                cell = ws.cell(row=current_row, column=2, value=sheet_info["month"])
                cell.font = font
            
            # C列到BF列（3-58）：全部生成引用公式
            for col in self.FORMULA_COLUMNS:
                if not self._is_merged_cell(ws, current_row, col):
                    formula = self.generate_formula(sheet_name, summary_row, col)
                    cell = ws.cell(row=current_row, column=col, value=formula)
                    cell.font = font
                    cell.number_format = number_format
            
            print(f"  ✅ 生成汇总行: {sheet_info['bank']} - {sheet_info['month']} (引用 {sheet_name} 第 {summary_row} 行)")
        
        # ===== 设置汇总表所有单元格字体和格式（跳过合并单元格） =====
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if not self._is_merged_cell(ws, row, col):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.font = font
                        if col >= 3:
                            cell.number_format = number_format
        
        # 保存
        wb.save(str(template_path))
        print(f"✅ 汇总表已更新: {template_path}")
        print(f"📝 已应用字体: 微软雅黑 8号")
        print(f"📝 已应用数字格式: 0 显示为 '-'")
        
        return str(template_path)