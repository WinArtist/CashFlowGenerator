# src/create_blank_template.py
"""现金流量报表生成器 - 根据明细账数据量生成模板 - 优化版"""

import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import copy
import re
import pandas as pd

__all__ = [
    'create_template_from_detail',
    'learn_protected_rows_styles',
    'copy_row_style',
    'update_data_formula',
    'set_balance_formula',
    'generate_protected_rows',
    'PROTECTED_ROWS_START',
    'PROTECTED_ROWS_END',
]

SKIP_VOUCHERS = {"期初余额", "本期合计", "本年累计", ""}  # 使用集合提高查找速度
PROTECTED_ROWS_START = 7
PROTECTED_ROWS_END = 13


def find_data_start_row(df):
    """查找数据起始行"""
    for idx, row in df.iterrows():
        for col in range(min(len(row), 30)):
            val = str(row.iloc[col]) if pd.notna(row.iloc[col]) else ''
            if '凭证字号' in val:
                return idx + 1
            if val.startswith('记-'):
                return idx
    return 0


def find_column_indices(df, start_row):
    """查找列索引"""
    header_row = df.iloc[start_row - 1] if start_row > 0 else df.iloc[0]
    col_index = {}
    for i, val in enumerate(header_row):
        if pd.notna(val):
            val_str = str(val).strip()
            if '日期' in val_str:
                col_index['date'] = i
            elif '凭证字号' in val_str:
                col_index['voucher'] = i
            elif '摘要' in val_str:
                col_index['desc'] = i
            elif '对方科目' in val_str:
                col_index['contra'] = i
            elif '借方' in val_str:
                col_index['debit'] = i
            elif '贷方' in val_str:
                col_index['credit'] = i
    return col_index


def count_real_transactions(file_path: str) -> int:
    """统计明细账中的真实交易数量 - 优化版"""
    print(f"\n📊 扫描明细账文件: {file_path}")
    
    try:
        excel_file = pd.ExcelFile(file_path)
        total_transactions = 0
        
        for sheet_name in excel_file.sheet_names:
            if '明细账' not in sheet_name and '明细' not in sheet_name:
                continue
            
            print(f"   检查工作表: {sheet_name}")
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str)
            start_row = find_data_start_row(df)
            col_index = find_column_indices(df, start_row)
            
            if not col_index or 'voucher' not in col_index:
                continue
            
            voucher_idx = col_index['voucher']
            debit_idx = col_index.get('debit')
            credit_idx = col_index.get('credit')
            
            sheet_count = 0
            for idx in range(start_row, len(df)):
                row = df.iloc[idx]
                voucher = str(row[voucher_idx]).strip() if pd.notna(row[voucher_idx]) else ''
                
                if not voucher or voucher in SKIP_VOUCHERS:
                    continue
                
                if debit_idx is not None and credit_idx is not None:
                    debit = 0
                    credit = 0
                    try:
                        debit = float(row[debit_idx]) if pd.notna(row[debit_idx]) else 0
                        credit = float(row[credit_idx]) if pd.notna(row[credit_idx]) else 0
                    except:
                        pass
                    if debit == 0 and credit == 0:
                        continue
                
                sheet_count += 1
            
            print(f"     真实交易数: {sheet_count}")
            total_transactions += sheet_count
        
        print(f"\n📈 总计真实交易数: {total_transactions}")
        return total_transactions
        
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return 0


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    try:
        if source_cell.has_style:
            if source_cell.font:
                target_cell.font = copy.copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy.copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy.copy(source_cell.alignment)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        return True
    except:
        return False


def copy_row_style(ws, source_row: int, target_row: int):
    """复制整行的样式"""
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(row=source_row, column=col), ws.cell(row=target_row, column=col))


def get_row_style_template(ws, row: int) -> dict:
    """获取一行的样式模板"""
    template = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        template[col] = {
            'font': copy.copy(cell.font) if cell.font else None,
            'fill': copy.copy(cell.fill) if cell.fill else None,
            'border': copy.copy(cell.border) if cell.border else None,
            'alignment': copy.copy(cell.alignment) if cell.alignment else None,
            'number_format': cell.number_format,
        }
    return template


def apply_style_template(ws, target_row: int, template: dict):
    """应用样式模板到目标行"""
    for col, style in template.items():
        cell = ws.cell(row=target_row, column=col)
        if style['font']:
            cell.font = style['font']
        if style['fill']:
            cell.fill = style['fill']
        if style['border']:
            cell.border = style['border']
        if style['alignment']:
            cell.alignment = style['alignment']
        if style['number_format']:
            cell.number_format = style['number_format']


def update_data_formula(formula: str, target_row: int) -> str:
    """更新数据行公式"""
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        if row_num == 4:
            return f"{col_letter}{row_num}"
        return f"{col_letter}{target_row}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)


def set_balance_formula(ws, row: int):
    """设置单行余额公式"""
    ws.cell(row=row, column=18, value=f"=R{row-1}+SUM(N{row}:Q{row})-SUM(J{row}:M{row})")


def learn_protected_rows_styles(ws, start_row: int, end_row: int) -> dict:
    """学习原始表保护行的样式"""
    print(f"\n📋 学习原始表第 {start_row} - {end_row} 行保护行样式...")
    
    result = {
        'row_styles': [],
        'merged_cells': [],
        'row_count': end_row - start_row + 1,
        'start_row': start_row,
        'end_row': end_row,
        'row_values': [],
        'row_formulas': [],
    }
    
    for row in range(start_row, end_row + 1):
        result['row_styles'].append(get_row_style_template(ws, row))
        
        row_values = {}
        row_formulas = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_formulas[col] = cell.value
                else:
                    row_values[col] = cell.value
        result['row_values'].append(row_values)
        result['row_formulas'].append(row_formulas)
    
    if result['row_formulas']:
        print(f"   合计行公式数: {len(result['row_formulas'][0])}")
    
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            result['merged_cells'].append({
                'min_row': merged_range.min_row - start_row,
                'max_row': merged_range.max_row - start_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
            })
    
    print(f"  已学习 {len(result['row_styles'])} 行样式")
    return result


def update_protected_formula(formula: str, source_row: int, target_row: int, data_end_row: int) -> str:
    """更新保护行合计行公式中的行号"""
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_range(match):
        col1, row1, col2, row2 = match.groups()
        r1, r2 = int(row1), int(row2)
        if row1 == '5' and 5 <= r2 <= 22:
            return f"{col1}{row1}:{col2}{data_end_row}"
        if 23 <= r1 <= 29 and 23 <= r2 <= 29:
            offset1, offset2 = r1 - 23, r2 - 23
            return f"{col1}{target_row + offset1}:{col2}{target_row + offset2}"
        return match.group(0)
    
    formula = re.sub(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', replace_range, formula)
    
    def replace_cell(match):
        col_letter, row_num_str = match.groups()
        row_num = int(row_num_str)
        if row_num == 4:
            return f"{col_letter}{row_num}"
        if 5 <= row_num <= 22:
            new_row = min(5 + (row_num - 5), data_end_row)
            return f"{col_letter}{new_row}"
        if 23 <= row_num <= 29:
            return f"{col_letter}{target_row + (row_num - 23)}"
        return f"{col_letter}{row_num}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)


def generate_protected_rows(ws, protected_data: dict, target_start_row: int, data_end_row: int):
    """在目标位置生成保护行"""
    print(f"\n📋 生成保护行（第 {target_start_row} 行开始）...")
    
    row_count = protected_data['row_count']
    
    for i in range(row_count):
        ws.insert_rows(target_start_row + i)
    
    for i in range(row_count):
        dst_row = target_start_row + i
        is_total1_row = (i == 0)
        
        apply_style_template(ws, dst_row, protected_data['row_styles'][i])
        
        for col, value in protected_data['row_values'][i].items():
            ws.cell(row=dst_row, column=col, value=value)
        
        if is_total1_row:
            for col, formula in protected_data['row_formulas'][i].items():
                ws.cell(row=dst_row, column=col, value=update_protected_formula(
                    formula, protected_data['start_row'] + i, dst_row, data_end_row
                ))
        else:
            for col in protected_data['row_formulas'][i].keys():
                ws.cell(row=dst_row, column=col, value=None)
    
    for merged in protected_data['merged_cells']:
        start_row = target_start_row + merged['min_row']
        end_row = target_start_row + merged['max_row']
        range_string = f"{get_column_letter(merged['min_col'])}{start_row}:{get_column_letter(merged['max_col'])}{end_row}"
        try:
            ws.merge_cells(range_string)
        except:
            pass
    
    print(f"  已生成 {row_count} 行保护行")


def create_template_from_detail(
    detail_path: str, 
    template_path: str, 
    output_dir: str,
    bank_name: str,
    period: str,
    output_filename: str,
    sheet_name: str
) -> str:
    """
    根据明细账数据量生成模板 - 支持追加到已有文件
    """
    detail_path = Path(detail_path)
    template_path = Path(template_path)
    
    if not detail_path.exists():
        print(f"❌ 明细账文件不存在: {detail_path}")
        return None
    
    if not template_path.exists():
        print(f"❌ 模板文件不存在: {template_path}")
        return None
    
    if not all([bank_name, period, output_filename, sheet_name]):
        print("❌ 必要参数缺失")
        return None
    
    data_rows = count_real_transactions(str(detail_path))
    if data_rows == 0:
        print("❌ 没有找到有效的交易数据")
        return None
    
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_filename
    
    # ===== 判断文件是否存在 =====
    file_exists = output_path.exists()
    
    # ===== 加载模板工作簿 =====
    template_wb = load_workbook(template_path)
    if '原始表' not in template_wb.sheetnames:
        print(f"❌ 模板文件中未找到 '原始表'")
        return None
    
    template_ws = template_wb['原始表']
    
    # ===== 加载或创建工作簿 =====
    if file_exists:
        print(f"\n📁 文件已存在: {output_path}")
        print(f"📄 将在该文件中添加新Sheet: {sheet_name}")
        wb = load_workbook(output_path)
        
        # 检查同名Sheet是否存在，存在则删除
        if sheet_name in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' 已存在，将被覆盖")
            std = wb[sheet_name]
            wb.remove(std)
    else:
        print(f"\n📁 创建新文件: {output_path}")
        # 复制模板作为基础
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        
        # 删除模板自带的"原始表"（我们要用自定义的）
        if '原始表' in wb.sheetnames:
            wb.remove(wb['原始表'])
        # 删除"汇总表"（保留模板中的结构）
        # 注意：如果模板有"汇总表"，保留它
        # 如果有其他不需要的Sheet，可以删除
    
    # ===== 创建新Sheet =====
    new_ws = wb.create_sheet(title="原始表_临时")
    
    # 复制模板"原始表"的内容到新Sheet
    max_row, max_col = template_ws.max_row, template_ws.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            src = template_ws.cell(row=row, column=col)
            dst = new_ws.cell(row=row, column=col)
            dst.value = src.value
            if src.has_style:
                if src.font:
                    dst.font = copy.copy(src.font)
                if src.fill:
                    dst.fill = copy.copy(src.fill)
                if src.border:
                    dst.border = copy.copy(src.border)
                if src.alignment:
                    dst.alignment = copy.copy(src.alignment)
                if src.number_format:
                    dst.number_format = src.number_format
    
    # 复制合并单元格
    for merged_range in template_ws.merged_cells.ranges:
        try:
            new_ws.merge_cells(str(merged_range))
        except:
            pass
    
    # 复制列宽和行高
    for col in template_ws.column_dimensions:
        new_ws.column_dimensions[col].width = template_ws.column_dimensions[col].width
    for row in template_ws.row_dimensions:
        new_ws.row_dimensions[row].height = template_ws.row_dimensions[row].height
    
    ws = new_ws
    
    # ===== 处理数据 =====
    style_data = learn_protected_rows_styles(ws, PROTECTED_ROWS_START, PROTECTED_ROWS_END)
    
    # 删除原有行
    for row in range(PROTECTED_ROWS_END, PROTECTED_ROWS_START - 1, -1):
        ws.delete_rows(row)
    for row in range(22, 5, -1):
        ws.delete_rows(row)
    
    # 插入数据行
    new_data_start = 5
    new_data_end = new_data_start + data_rows - 1
    new_protected_start = new_data_end + 1
    
    rows_to_insert = data_rows - 1
    if rows_to_insert > 0:
        ws.insert_rows(new_data_start + 1, rows_to_insert)
    
    # 复制样式
    template_style_row = 5
    for target_row in range(new_data_start + 1, new_data_end + 1):
        copy_row_style(ws, template_style_row, target_row)
    
    # 收集公式模板
    template_formulas = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=5, column=col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') and col != 18:
            template_formulas[col] = cell.value
    
    # 生成公式
    for target_row in range(new_data_start, new_data_end + 1):
        for col, formula in template_formulas.items():
            ws.cell(row=target_row, column=col, value=update_data_formula(formula, target_row))
        set_balance_formula(ws, target_row)
    
    # 生成保护行
    generate_protected_rows(ws, style_data, new_protected_start, new_data_end)
    
    generate_summary_formula(ws, new_data_start, new_data_end, new_protected_start)

    # ===== 重命名Sheet =====
    if "原始表_临时" in wb.sheetnames:
        wb["原始表_临时"].title = sheet_name
    
    # ===== 保存（保留所有已有Sheet） =====
    wb.save(str(output_path))
    print(f"\n✅ 模板生成完成!")
    print(f"📁 输出文件: {output_path}")
    print(f"📄 工作表名称: {sheet_name}")
    print(f"📊 数据行数: {data_rows}")
    print(f"📋 当前文件中的所有Sheet: {wb.sheetnames}")
    
    return str(output_path)

def generate_summary_formula(ws, data_start_row: int, data_end_row: int, summary_row: int):
    """
    生成合计行公式
    """
    # ===== 核心余额公式 (R列) =====
    ws.cell(row=summary_row, column=18, value=f"=R4+SUM(O{data_end_row + 1}:Q{data_end_row + 1})-SUM(K{data_end_row + 1}:N{data_end_row + 1})-J{data_end_row + 1}")