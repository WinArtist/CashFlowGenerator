# src/services/direct_sheet_filler.py
"""直接填入工作表服务 - 优化版"""

from typing import Optional
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
import re

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import Config


class JianhangQ2Filler:
    """建行Q2工作表填充器 - 优化版"""
    
    def __init__(self, config: Config):
        self.config = config
        self.decimal_places = config.data.decimal_places
        self.COL_INDEX = config.data.classification.column_mapping.copy()
        self.COL_INDEX.update({
            '交易时间': 2,
            '客户供应商': 19,
            '摘要': 21,
            '余额': 18,  # R列
        })
    
    def fill_quarter_data(self, template_path: str, report_data, quarter: str, source_file: str, sheet_name: str, opening_balance: Optional[Decimal] = None) -> Optional[str]:
        from openpyxl import load_workbook
        
        if not template_path:
            raise ValueError("模板文件路径未设置")
        
        template_path = Path(template_path)
        if not template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_path}")
        
        if not quarter or not sheet_name:
            raise ValueError("季度和工作表名称不能为空")
        
        wb = load_workbook(template_path)
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"未找到工作表: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # ===== 填入期初余额到 R4 =====
        if opening_balance is not None:
            try:
                ws.cell(row=4, column=18, value=float(opening_balance))
                print(f"  ✅ 期初余额已填入 R4: {opening_balance:.2f}")
            except Exception as e:
                print(f"  ❌ 填入期初余额失败: {e}")
        else:
            print("  ⚠️ 未找到期初余额，R4 保持为空")
        
        start_row = 5
        transactions = getattr(report_data, '_transactions', [])
        
        if transactions:
            filled_count = 0
            for idx, trans in enumerate(transactions):
                row = start_row + idx
                try:
                    self._fill_transaction_row(ws, row, trans)
                    filled_count += 1
                except Exception as e:
                    continue
            
            print(f"  成功填入 {filled_count}/{len(transactions)} 笔交易")
        else:
            print("  警告: 没有原始交易数据")
        
        wb.save(str(template_path))
        print(f"✅ 文件已保存: {template_path}")
        
        return str(template_path)
    
    def _set_cell_value(self, ws, row: int, col: int, value):
        try:
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if isinstance(value, (datetime, date)):
                cell.number_format = 'yyyy/m/d'
            return True
        except Exception:
            return False
    
    def _should_fill_contra(self, contra_subject: str, summary: str = "") -> bool:
        """判断是否应该填充客户/供应商列（白名单模式）"""
        if not contra_subject:
            return False
        
        contra_subject = str(contra_subject)
        classification = self.config.data.classification
        allow_prefixes = getattr(classification, 'allow_prefixes', ["1122", "1123", "2202", "2203"])
        
        for prefix in allow_prefixes:
            if contra_subject.startswith(prefix):
                return True
        
        return False
    
    def _extract_company_name(self, contra_subject: str) -> str:
        """从对方科目中提取公司名称"""
        if not contra_subject:
            return ""
        
        contra_subject = str(contra_subject)
        
        # ===== 1. 标准格式：应收账款_人民币户_苏州登临科技股份有限公司 =====
        match = re.search(r'(?:应收|应付)(?:账款|票据)?_?(?:人民币户)?_?([^\s_]+(?:公司|有限|集团|中心|厂|店|银行))', contra_subject)
        if match:
            return match.group(1)
        
        # ===== 2. 简单格式：应收账款_公司名 =====
        match = re.search(r'(?:应收|应付)(?:账款|票据)?_([^\s_]+(?:公司|有限|集团|中心|厂|店|银行))', contra_subject)
        if match:
            return match.group(1)
        
        # ===== 3. 其他应付款_徐芳 / 其他应收款_朱晓福 =====
        match = re.search(r'其他应(?:收|付)款_([^\s_]+)', contra_subject)
        if match:
            return match.group(1)
        
        # ===== 4. 直接匹配公司后缀 =====
        match = re.search(r'([^\s_]{2,20}(?:公司|有限|集团|中心|厂|店|银行))', contra_subject)
        if match:
            return match.group(1)
        
        # ===== 5. 按_分割取最后一段 =====
        parts = contra_subject.split('_')
        if len(parts) >= 2:
            last_part = parts[-1].strip()
            if last_part and len(last_part) >= 2 and not last_part.isdigit():
                if re.search(r'[\u4e00-\u9fa5]', last_part):
                    return last_part
        
        # ===== 6. 提取连续的中文字符（2-20个） =====
        match = re.search(r'([\u4e00-\u9fa5]{2,20})', contra_subject)
        if match:
            return match.group(1)
        
        return ""
    
    def _fill_transaction_row(self, ws, row: int, trans):
        """填充单笔交易到指定行"""
        # 交易时间 - B列
        if hasattr(trans, 'date') and trans.date:
            if isinstance(trans.date, datetime):
                date_value = trans.date.date()
            else:
                date_value = trans.date
            self._set_cell_value(ws, row, self.COL_INDEX.get('交易时间', 2), date_value)
        
        # ===== 对方科目 - S列（客户/供应商）- 白名单模式 =====
        contra_subject = getattr(trans, 'contra_subject', "")
        description = getattr(trans, 'description', "")
        
        if self._should_fill_contra(contra_subject, description):
            company_name = self._extract_company_name(contra_subject)
            if company_name:
                self._set_cell_value(ws, row, self.COL_INDEX.get('客户供应商', 19), company_name)
            else:
                self._set_cell_value(ws, row, self.COL_INDEX.get('客户供应商', 19), contra_subject)
        
        # 摘要 - U列
        if description:
            desc = str(description)
            if len(desc) > 50:
                desc = desc[:47] + "..."
            self._set_cell_value(ws, row, self.COL_INDEX.get('摘要', 21), desc)
        
        # ===== 余额 - R列 =====
        if hasattr(trans, 'balance') and trans.balance is not None:
            self._set_cell_value(ws, row, self.COL_INDEX.get('余额', 18), float(trans.balance))
        
        # ===== 收入分类 - 所有收入都填入"其他收入"列 =====
        if trans.is_income:
            # 所有收入都放入"其他收入"列（第17列）
            self._set_cell_value(ws, row, self.COL_INDEX.get('其他收入', 17), float(trans.amount))
        
        # ===== 支出分类 =====
        else:
            expense_category = getattr(trans, 'expense_category', None)
            if expense_category:
                col = self.COL_INDEX.get(expense_category)
                if col:
                    self._set_cell_value(ws, row, col, float(trans.amount))