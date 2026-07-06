# main.py
"""现金流量报表生成器 - 根据明细账数据量生成建行Q2模板"""

import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import copy
import re
import pandas as pd


# ============================================================
# 配置
# ============================================================

SKIP_VOUCHERS = ["期初余额", "本期合计", "本年累计", ""]

# 原始表中需要学习的保护行范围（从第23行开始）
PROTECTED_ROWS_START = 23
PROTECTED_ROWS_END = 29


def find_data_start_row(df):
    """查找数据起始行"""
    for idx in range(len(df)):
        row = df.iloc[idx]
        for col in range(len(row)):
            val = str(row.iloc[col]) if pd.notna(row.iloc[col]) else ''
            if '凭证字号' in val:
                return idx + 1
            if val.startswith('记-'):
                return idx
    return 0


def find_column_indices(df, start_row):
    """查找列索引"""
    header_row = df.iloc[start_row - 1] if start_row > 0 else df.iloc[0]
    
    col_index = {}
    for i, val in enumerate(header_row):
        if pd.notna(val):
            val_str = str(val).strip()
            if '日期' in val_str:
                col_index['date'] = i
            elif '凭证字号' in val_str:
                col_index['voucher'] = i
            elif '摘要' in val_str:
                col_index['desc'] = i
            elif '对方科目' in val_str:
                col_index['contra'] = i
            elif '借方' in val_str:
                col_index['debit'] = i
            elif '贷方' in val_str:
                col_index['credit'] = i
    
    return col_index


def count_real_transactions(file_path: str) -> int:
    """统计明细账中的真实交易数量"""
    print(f"\n📊 扫描明细账文件: {file_path}")
    
    try:
        excel_file = pd.ExcelFile(file_path)
        total_transactions = 0
        
        for sheet_name in excel_file.sheet_names:
            if '明细账' in sheet_name or '明细' in sheet_name:
                print(f"   检查工作表: {sheet_name}")
                
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                start_row = find_data_start_row(df)
                col_index = find_column_indices(df, start_row)
                
                if not col_index or 'voucher' not in col_index:
                    continue
                
                voucher_idx = col_index['voucher']
                debit_idx = col_index.get('debit')
                credit_idx = col_index.get('credit')
                
                sheet_count = 0
                for idx in range(start_row, len(df)):
                    row = df.iloc[idx]
                    voucher = row[voucher_idx]
                    if pd.isna(voucher):
                        voucher = ''
                    else:
                        voucher = str(voucher).strip()
                    
                    if not voucher or voucher in SKIP_VOUCHERS:
                        continue
                    
                    if debit_idx is not None and credit_idx is not None:
                        debit = row[debit_idx] if pd.notna(row[debit_idx]) else 0
                        credit = row[credit_idx] if pd.notna(row[credit_idx]) else 0
                        try:
                            debit = float(debit) if debit else 0
                            credit = float(credit) if credit else 0
                        except:
                            debit = 0
                            credit = 0
                        if debit == 0 and credit == 0:
                            continue
                    
                    sheet_count += 1
                
                print(f"     真实交易数: {sheet_count}")
                total_transactions += sheet_count
        
        print(f"\n📈 总计真实交易数: {total_transactions}")
        return total_transactions
        
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return 0


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    try:
        if source_cell.has_style:
            if source_cell.font:
                target_cell.font = copy.copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy.copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy.copy(source_cell.alignment)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        return True
    except:
        return False


def copy_row_style(ws, source_row: int, target_row: int):
    """复制整行的样式（不复制值和公式）"""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        copy_cell_style(source_cell, target_cell)


def get_row_style_template(ws, row: int) -> dict:
    """获取一行的样式模板"""
    template = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        template[col] = {
            'font': copy.copy(cell.font) if cell.font else None,
            'fill': copy.copy(cell.fill) if cell.fill else None,
            'border': copy.copy(cell.border) if cell.border else None,
            'alignment': copy.copy(cell.alignment) if cell.alignment else None,
            'number_format': cell.number_format,
        }
    return template


def apply_style_template(ws, target_row: int, template: dict):
    """应用样式模板到目标行"""
    for col, style in template.items():
        cell = ws.cell(row=target_row, column=col)
        if style['font']:
            cell.font = style['font']
        if style['fill']:
            cell.fill = style['fill']
        if style['border']:
            cell.border = style['border']
        if style['alignment']:
            cell.alignment = style['alignment']
        if style['number_format']:
            cell.number_format = style['number_format']


def update_data_formula(formula: str, target_row: int) -> str:
    """
    更新数据行公式 - 将公式中的所有行号改为目标行号
    第4行（期初余额）保持不变
    """
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        if row_num == 4:
            return f"{col_letter}{row_num}"
        return f"{col_letter}{target_row}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)






def set_balance_formula(ws, row: int):
    """设置单行余额公式，逐行递增引用上一行"""
    formula = f"=R{row-1}+SUM(N{row}:Q{row})-SUM(J{row}:M{row})"
    ws.cell(row=row, column=18, value=formula)

def learn_protected_rows_styles(ws, start_row: int, end_row: int) -> dict:
    """学习原始表保护行的样式"""
    print(f"\n📋 学习原始表第 {start_row} - {end_row} 行保护行样式...")
    
    result = {
        'row_styles': [],
        'merged_cells': [],
        'row_count': end_row - start_row + 1,
        'start_row': start_row,
        'end_row': end_row,
        'row_values': [],
        'row_formulas': [],
    }
    
    for row in range(start_row, end_row + 1):
        style_template = get_row_style_template(ws, row)
        result['row_styles'].append(style_template)
        
        row_values = {}
        row_formulas = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_formulas[col] = cell.value
                else:
                    row_values[col] = cell.value
        result['row_values'].append(row_values)
        result['row_formulas'].append(row_formulas)
    
    # 打印合计行的公式，用于调试
    if result['row_formulas']:
        print(f"   合计行公式数: {len(result['row_formulas'][0])}")
        for col, formula in list(result['row_formulas'][0].items())[:10]:
            print(f"     第{col}列: {formula}")
    
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row >= start_row and 
            merged_range.max_row <= end_row):
            result['merged_cells'].append({
                'min_row': merged_range.min_row - start_row,
                'max_row': merged_range.max_row - start_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
            })
    
    print(f"  已学习 {len(result['row_styles'])} 行样式")
    print(f"  已学习 {len(result['merged_cells'])} 个合并单元格")
    
    return result

def update_protected_formula(formula: str, source_row: int, target_row: int, data_end_row: int) -> str:
    """
    更新保护行合计行公式中的行号（仅用于第23行和第29行）
    将公式中的行号范围（如 D5:D22）更新为 D5:D{data_end_row}
    """
    if not formula or not formula.startswith('='):
        return formula
    
    # 如果是 SUM 范围公式，如 =SUM(D5:D22)
    # 将结束行号改为 data_end_row
    def replace_range(match):
        col1 = match.group(1)
        row1 = match.group(2)
        col2 = match.group(3)
        row2 = match.group(4)
        
        # 如果起始行是5，结束行在5-22之间，更新结束行为 data_end_row
        if row1 == '5' and 5 <= int(row2) <= 22:
            return f"{col1}{row1}:{col2}{data_end_row}"
        # 如果是保护行范围（23-29），根据偏移量调整
        if 23 <= int(row1) <= 29 and 23 <= int(row2) <= 29:
            offset = int(row1) - 23
            new_row1 = target_row + offset
            offset2 = int(row2) - 23
            new_row2 = target_row + offset2
            return f"{col1}{new_row1}:{col2}{new_row2}"
        return match.group(0)
    
    # 先处理范围引用
    formula = re.sub(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', replace_range, formula)
    
    # 再处理单个单元格引用（如 R4, N5 等）
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        
        # 第4行是期初余额，保持不变
        if row_num == 4:
            return f"{col_letter}{row_num}"
        
        # 如果是数据行（5-22），映射到对应的数据行
        if 5 <= row_num <= 22:
            offset = row_num - 5
            new_row_num = 5 + offset
            if new_row_num > data_end_row:
                new_row_num = data_end_row
            return f"{col_letter}{new_row_num}"
        
        # 如果是保护行（23-29），根据偏移量调整
        if 23 <= row_num <= 29:
            offset = row_num - 23
            new_row_num = target_row + offset
            return f"{col_letter}{new_row_num}"
        
        # 其他行号保持不变
        return f"{col_letter}{row_num}"
    
    formula = re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)
    
    return formula


def update_data_style_formula(formula: str, target_row: int) -> str:
    """
    更新保护行中数据样式行（第25、26行）的公式
    将公式中的行号替换为目标行号（引用自身）
    - 范围引用（如 SUM(O25:Q25)）中的行号 -> target_row
    - R列引用（如 R24）中的行号 -> target_row - 1（上一行）
    """
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        
        # 第4行是期初余额，保持不变
        if row_num == 4:
            return f"{col_letter}{row_num}"
        
        # 如果是 R 列（余额列），引用上一行
        if col_letter == 'R':
            return f"{col_letter}{target_row - 1}"
        
        # 其他行号替换为目标行号
        return f"{col_letter}{target_row}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)

def generate_protected_rows(ws, protected_data: dict, target_start_row: int, data_end_row: int):
    """在目标位置生成保护行"""
    print(f"\n📋 生成保护行（第 {target_start_row} 行开始）...")
    
    row_count = protected_data['row_count']
    
    # 插入空行
    for i in range(row_count):
        ws.insert_rows(target_start_row + i)
    
    for i in range(row_count):
        dst_row = target_start_row + i
        
        # 只有第一个合计行（i == 0，原23行）保留公式
        is_total1_row = (i == 0)
        
        # 应用样式
        style_template = protected_data['row_styles'][i]
        apply_style_template(ws, dst_row, style_template)
        
        # 设置非公式值
        for col, value in protected_data['row_values'][i].items():
            ws.cell(row=dst_row, column=col, value=value)
        
        # 只有第一个合计行处理公式，其他行全部清除公式
        if is_total1_row:
            for col, formula in protected_data['row_formulas'][i].items():
                new_formula = update_protected_formula(
                    formula,
                    protected_data['start_row'] + i,
                    dst_row,
                    data_end_row
                )
                ws.cell(row=dst_row, column=col, value=new_formula)
        else:
            # 其他行：清除所有公式，只保留文本值
            for col in protected_data['row_formulas'][i].keys():
                ws.cell(row=dst_row, column=col, value=None)
    
    # 创建合并单元格
    for merged in protected_data['merged_cells']:
        start_row = target_start_row + merged['min_row']
        end_row = target_start_row + merged['max_row']
        start_col = merged['min_col']
        end_col = merged['max_col']
        range_string = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        try:
            ws.merge_cells(range_string)
        except:
            pass
    
    print(f"  已生成 {row_count} 行保护行")
    print(f"  保留第1行（合计行）公式，其他行公式已清除")

def update_protected_balance_formula(formula: str, target_row: int) -> str:
    """
    更新保护行中第28行（理财产品期末余额）的公式
    =R27+SUM(O28:Q28)-SUM(J28:N28)
    -> =R{target_row-1}+SUM(O{target_row}:Q{target_row})-SUM(J{target_row}:N{target_row})
    """
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        
        # 如果是 R 列（余额列），引用上一行
        if col_letter == 'R':
            return f"{col_letter}{target_row - 1}"
        
        # 其他行号替换为目标行号
        return f"{col_letter}{target_row}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)


def update_protected_total2_formula(formula: str, total1_row: int, total2_row: int) -> str:
    """
    更新保护行中第29行（第二个合计行）的公式
    =D28+D23 -> =D{total2_row-1}+D{total1_row}
    =E28+E23 -> =E{total2_row-1}+E{total1_row}
    """
    if not formula or not formula.startswith('='):
        return formula
    
    def replace_cell(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        
        # 如果是28行，替换为 total2_row - 1（上一行，即理财产品期末余额行）
        if row_num == 28:
            return f"{col_letter}{total2_row - 1}"
        # 如果是23行，替换为 total1_row（第一个合计行）
        if row_num == 23:
            return f"{col_letter}{total1_row}"
        # 其他行号保持不变
        return f"{col_letter}{row_num}"
    
    return re.sub(r'([A-Z]+)(\d+)', replace_cell, formula)

def create_template_from_detail(detail_path: str, template_path: str, output_dir: str = None) -> str:
    """根据明细账数据量生成建行Q2模板"""
    detail_path = Path(detail_path)
    template_path = Path(template_path)
    
    if not detail_path.exists():
        print(f"❌ 明细账文件不存在: {detail_path}")
        return None
    
    if not template_path.exists():
        print(f"❌ 模板文件不存在: {template_path}")
        return None
    
    data_rows = count_real_transactions(str(detail_path))
    
    if data_rows == 0:
        print("❌ 没有找到有效的交易数据")
        return None
    
    if output_dir is None:
        output_dir = Path(__file__).parent / "output"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"建行Q2_模板_{data_rows}行_{timestamp}.xlsx"
    
    print(f"\n📁 复制模板文件...")
    shutil.copy2(template_path, output_path)
    
    wb = load_workbook(output_path)
    
    if '原始表' not in wb.sheetnames:
        print(f"❌ 未找到 '原始表' 工作表")
        return None
    
    ws = wb['原始表']
    print(f"✅ 找到 '原始表' 工作表")
    
    # ============================================================
    # 步骤1: 学习原始表的保护行样式（必须在删除之前）
    # ============================================================
    print(f"\n{'='*50}")
    print("步骤1: 学习原始表保护行样式（第23-29行）...")
    style_data = learn_protected_rows_styles(ws, PROTECTED_ROWS_START, PROTECTED_ROWS_END)
    
    # ============================================================
    # 步骤2: 删除所有不需要的行（从下往上删，避免行号变化）
    # ============================================================
    print(f"\n步骤2: 删除所有不需要的行...")
    
    # 先删除保护行（23-29行）
    for row in range(PROTECTED_ROWS_END, PROTECTED_ROWS_START - 1, -1):
        ws.delete_rows(row)
    print(f"  已删除原始表保护行（第{PROTECTED_ROWS_START}-{PROTECTED_ROWS_END}行）")
    
    # 再删除数据行（6-22行），注意此时行号已经变化，保护行删除后，数据行仍然是6-22行
    for row in range(22, 5, -1):
        ws.delete_rows(row)
    print(f"  已删除原有数据行（第6-22行）")
    
    # ============================================================
    # 步骤3: 插入数据行
    # ============================================================
    new_data_start = 6
    new_data_end = new_data_start + data_rows - 1
    new_protected_start = new_data_end + 1
    
    print(f"\n步骤3: 插入数据行...")
    print(f"  数据行范围: 第 {new_data_start} - {new_data_end} 行")
    print(f"  保护行起始: 第 {new_protected_start} 行")
    
    rows_to_insert = data_rows - 1
    if rows_to_insert > 0:
        ws.insert_rows(new_data_start + 1, rows_to_insert)
        print(f"  已插入 {rows_to_insert} 行")
    
    # ============================================================
    # 步骤4: 复制样式到数据行
    # ============================================================
    print(f"\n步骤4: 复制样式到数据行...")
    template_style_row = 5
    for target_row in range(new_data_start, new_data_end + 1):
        copy_row_style(ws, template_style_row, target_row)
    print(f"  已复制样式到 {data_rows} 行")
    
    # ============================================================
    # 步骤5: 收集第5行的所有公式模板
    # ============================================================
    print(f"\n步骤5: 收集模板行公式...")
    template_formulas = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=5, column=col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            template_formulas[col] = cell.value
    
    print(f"   模板行包含 {len(template_formulas)} 个公式（不含余额列）")
    
    # ============================================================
    # 步骤6: 为每一行生成公式
    # ============================================================
    print(f"\n步骤6: 生成数据行公式...")
    for target_row in range(new_data_start, new_data_end + 1):
        for col, formula in template_formulas.items():
            if col == 18:  # 余额列单独处理
                continue
            new_formula = update_data_formula(formula, target_row)
            ws.cell(row=target_row, column=col, value=new_formula)
        
        set_balance_formula(ws, target_row)
    
    print(f"  已设置 {data_rows} 行公式")
    
    # 验证
    print(f"\n   验证数据行公式（前3行）:")
    for row in [new_data_start, min(new_data_start + 1, new_data_end)]:
        if row <= new_data_end:
            c_val = ws.cell(row=row, column=3).value
            r_val = ws.cell(row=row, column=18).value
            print(f"     第{row}行 C列: {c_val}")
            print(f"     第{row}行 R列: {r_val}")
    
    # ============================================================
    # 步骤7: 生成保护行
    # ============================================================
    print(f"\n步骤7: 生成保护行...")
    generate_protected_rows(ws, style_data, new_protected_start, new_data_end)
    
    # 验证保护行合计行的公式
    print(f"\n   验证保护行合计行公式:")
    total_row = new_protected_start
    for col in [3, 4, 5]:  # C列、D列、E列
        val = ws.cell(row=total_row, column=col).value
        print(f"     第{total_row}行 第{col}列: {val}")
    
    # ============================================================
    # 打印汇总
    # ============================================================
    print(f"\n{'='*50}")
    print("✅ 模板生成完成!")
    print(f"{'='*50}")
    print(f"\n📁 输出文件: {output_path}")
    print(f"\n📊 统计信息:")
    print(f"   - 数据行数: {data_rows}")
    print(f"   - 数据行范围: 第 {new_data_start} - {new_data_end} 行")
    print(f"   - 保护行: 第 {new_protected_start} - {new_protected_start + style_data['row_count'] - 1} 行")
    
    wb.save(output_path)
    print(f"\n💾 文件已保存")
    
    return str(output_path)

def scan_input_directory() -> list:
    """扫描input目录下的所有Excel文件"""
    input_dir = Path(__file__).parent / 'input'
    
    if not input_dir.exists():
        print(f"目录不存在: {input_dir}")
        return []
    
    excel_files = [f for f in input_dir.glob("*.xlsx") if not f.name.startswith('~$')]
    excel_files += [f for f in input_dir.glob("*.xls") if not f.name.startswith('~$')]
    return sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True)


def display_files_table(files: list) -> None:
    """显示文件列表"""
    if not files:
        print("未找到任何Excel文件")
        return
    
    print("\n" + "=" * 80)
    print(f"{'序号':<6} {'文件名':<50} {'大小':<12}")
    print("=" * 80)
    
    for idx, file_path in enumerate(files, 1):
        size_kb = file_path.stat().st_size / 1024
        size_str = f"{size_kb:.1f} KB" if size_kb < 1024 else f"{size_kb/1024:.1f} MB"
        filename = file_path.name
        if len(filename) > 48:
            filename = filename[:45] + "..."
        print(f"{idx:<6} {filename:<50} {size_str:<12}")
    
    print("=" * 80)
    print(f"共找到 {len(files)} 个文件")


def main():
    """主函数"""
    print("=" * 60)
    print("    建行Q2模板生成器")
    print("=" * 60)
    
    input_dir = Path(__file__).parent / 'input'
    if not input_dir.exists():
        input_dir.mkdir(parents=True, exist_ok=True)
        print(f"\n📁 已创建 input 目录: {input_dir}")
        print(f"   请将明细账文件放入该目录")
        return 1
    
    files = scan_input_directory()
    
    if not files:
        print(f"\n❌ input目录下没有找到Excel文件")
        print(f"   请将明细账文件放入: {input_dir}")
        return 1
    
    display_files_table(files)
    
    if len(files) == 1:
        selected = files[0]
        print(f"\n自动选择: {selected.name}")
    else:
        choice = input(f"\n请选择明细账文件 (1-{len(files)}): ").strip()
        try:
            idx = int(choice) - 1
            if idx < 0 or idx >= len(files):
                print("无效选择")
                return 1
            selected = files[idx]
        except:
            print("无效选择")
            return 1
    
    template_path = Path(__file__).parent / 'templates' / '现金流原始表.xlsx'
    
    if not template_path.exists():
        print(f"\n❌ 模板文件不存在: {template_path}")
        print(f"   请将模板文件放入: {Path(__file__).parent / 'templates'}")
        return 1
    
    print(f"\n📁 明细账文件: {selected}")
    print(f"📁 模板文件: {template_path}")
    
    confirm = input(f"\n确认生成模板? (y/n): ").strip().lower()
    if confirm != 'y':
        print("已取消")
        return 0
    
    output_path = create_template_from_detail(str(selected), str(template_path))
    
    if output_path:
        print(f"\n{'='*60}")
        print("✅ 完成!")
        print(f"📁 {output_path}")
        print("=" * 60)
    
    return 0


if __name__ == "__main__":
    exit(main())